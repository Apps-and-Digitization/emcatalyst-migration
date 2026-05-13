"""
MCL Doctor Import - Handles large Excel files (400K+ rows)
Uses chunked processing with openpyxl in read-only mode for memory efficiency.
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks
from sqlalchemy.orm import Session
from sqlalchemy import text
import os
import tempfile
import logging

from app.db.base import get_db, SessionLocal
from app.api.deps import get_current_active_user, require_admin
from app.models.user import User

router = APIRouter(prefix="/import", tags=["import"])
logger = logging.getLogger(__name__)

# Column mapping: Excel header → DB column
COLUMN_MAP = {
    'Division': 'division',
    'Terriroty Name': 'territory_name',
    'Territory Name': 'territory_name',
    'Employee Code': 'employee_code',
    'User First Name': 'employee_code',  # fallback
    'UID Number': 'uid_number',
    'SBU Code': 'sbu_code',
    'Dr First Name': 'first_name',
    'Dr Middle Name': 'middle_name',
    'Dr Last Name': 'last_name',
    'Dr Gender': 'gender',
    'Dr Type': 'doctor_type',
    'Qualification': 'qualification',
    'Speciality': 'speciality',
    'City': 'city',
    'State': 'state',
    'Town Name': 'town_name',
    'Birthday': 'birthday',
    'Service Preference': 'service_preference',
    'Area Of Practice': 'area_of_practice',
    'Mobile': 'mobile_number',
    'Email': 'email',
}

# Import status tracking
import_status = {}


def process_excel_file(file_path: str, import_id: str):
    """Process Excel file in background - handles 400K+ rows"""
    import openpyxl

    import_status[import_id] = {"status": "processing", "total": 0, "processed": 0, "errors": 0}

    try:
        # Use read_only mode for memory efficiency with large files
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        # Get headers from first row
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(h).strip() if h else '' for h in row]
            break

        # Map headers to DB columns
        col_mapping = {}
        for i, header in enumerate(headers):
            if header in COLUMN_MAP:
                col_mapping[i] = COLUMN_MAP[header]

        if not col_mapping:
            import_status[import_id] = {"status": "error", "message": "No matching columns found in Excel"}
            return

        # Count total rows (approximate for large files)
        total_rows = ws.max_row - 1 if ws.max_row else 0
        import_status[import_id]["total"] = total_rows

        # Process in chunks
        CHUNK_SIZE = 1000
        chunk = []
        processed = 0
        errors = 0

        db = SessionLocal()
        try:
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                # Build record from row
                record = {}
                for col_idx, db_col in col_mapping.items():
                    if col_idx < len(row):
                        val = row[col_idx]
                        if val is not None:
                            record[db_col] = str(val).strip()

                if not record.get('first_name') and not record.get('last_name'):
                    continue  # Skip empty rows

                # Build full_name
                parts = [record.get('first_name', ''), record.get('middle_name', ''), record.get('last_name', '')]
                record['full_name'] = ' '.join(p for p in parts if p and p != '.')
                record['is_active'] = True

                chunk.append(record)

                if len(chunk) >= CHUNK_SIZE:
                    _insert_chunk(db, chunk)
                    processed += len(chunk)
                    import_status[import_id]["processed"] = processed
                    chunk = []

            # Insert remaining
            if chunk:
                _insert_chunk(db, chunk)
                processed += len(chunk)

            import_status[import_id] = {
                "status": "completed",
                "total": total_rows,
                "processed": processed,
                "errors": errors,
            }
        finally:
            db.close()
            wb.close()

    except Exception as e:
        logger.error(f"Import error: {e}")
        import_status[import_id] = {"status": "error", "message": str(e)}
    finally:
        # Clean up temp file
        try:
            os.unlink(file_path)
        except:
            pass


def _insert_chunk(db: Session, records: list):
    """Insert a chunk of records using bulk insert with upsert on uid_number"""
    from app.models.master import HcpDoctor, HcpDoctorDivision
    from app.models.user import Division

    for record in records:
        division_name = record.pop('division', None)
        uid = record.get('uid_number')

        if uid:
            # Update existing by UID
            existing = db.query(HcpDoctor).filter(HcpDoctor.uid_number == uid).first()
            if existing:
                for k, v in record.items():
                    if v and k != 'uid_number':
                        setattr(existing, k, v)
                # Add division association if not exists
                if division_name:
                    existing.division = division_name
                    div = db.query(Division).filter(Division.name.ilike(f"%{division_name}%")).first()
                    if div:
                        exists = db.query(HcpDoctorDivision).filter(
                            HcpDoctorDivision.hcp_doctor_id == existing.id,
                            HcpDoctorDivision.division_id == div.id
                        ).first()
                        if not exists:
                            db.add(HcpDoctorDivision(hcp_doctor_id=existing.id, division_id=div.id))
                continue

        # Insert new
        if division_name:
            record['division'] = division_name
        doc = HcpDoctor(**record)
        db.add(doc)
        db.flush()

        # Link to division
        if division_name:
            div = db.query(Division).filter(Division.name.ilike(f"%{division_name}%")).first()
            if div:
                db.add(HcpDoctorDivision(hcp_doctor_id=doc.id, division_id=div.id))

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        logger.error(f"Chunk insert error: {e}")


@router.post("/mcl")
async def import_mcl_excel(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin)
):
    """
    Import MCL doctors from Excel file.
    Handles 400K+ rows using background processing.
    Upserts based on UID Number.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Only .xlsx or .xls files are allowed")

    # Save to temp file
    temp_dir = tempfile.mkdtemp()
    temp_path = os.path.join(temp_dir, file.filename)
    with open(temp_path, "wb") as f:
        content = await file.read()
        f.write(content)

    # Generate import ID
    import_id = f"import_{os.urandom(4).hex()}"

    # Process in background
    background_tasks.add_task(process_excel_file, temp_path, import_id)

    return {
        "import_id": import_id,
        "message": "Import started in background. Check status with GET /import/status/{import_id}",
        "file_name": file.filename,
        "file_size_mb": round(len(content) / (1024 * 1024), 2),
    }


@router.get("/status/{import_id}")
def get_import_status(import_id: str, current_user: User = Depends(require_admin)):
    """Check the status of a background import"""
    if import_id not in import_status:
        raise HTTPException(404, "Import not found")
    return import_status[import_id]


@router.get("/mcl/template")
def get_mcl_template(current_user: User = Depends(require_admin)):
    """Return the expected column headers for MCL import"""
    return {
        "columns": list(COLUMN_MAP.keys()),
        "description": "Excel file must have these column headers in the first row",
    }
